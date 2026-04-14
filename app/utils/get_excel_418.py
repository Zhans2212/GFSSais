from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def rows_to_pdf(rows=None, date=None, fio=None, approved_by=None):
    buffer = BytesIO()

    font_regular = Path("app/static/fonts/times.ttf")
    font_bold = Path("app/static/fonts/timesbd.ttf")
    font_italic = Path("app/static/fonts/timesit.ttf")

    pdfmetrics.registerFont(TTFont("TimesNewRoman", str(font_regular)))
    pdfmetrics.registerFont(TTFont("TimesNewRoman-Bold", str(font_bold)))
    pdfmetrics.registerFont(TTFont("TimesNewRoman-Italic", str(font_italic)))

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()

    normal_style = ParagraphStyle(
        "NormalTNR",
        parent=styles["Normal"],
        fontName="TimesNewRoman",
        fontSize=11,
        leading=14,
        alignment=TA_LEFT,
    )

    italic_style = ParagraphStyle(
        "ItalicTNR",
        parent=styles["Normal"],
        fontName="TimesNewRoman-Italic",
        fontSize=11,
        leading=14,
        alignment=TA_LEFT,
    )

    title_style = ParagraphStyle(
        "TitleTNR",
        parent=styles["Title"],
        fontName="TimesNewRoman-Bold",
        fontSize=14,
        leading=18,
        alignment=TA_CENTER,
    )

    center_style = ParagraphStyle(
        "CenterTNR",
        parent=styles["Normal"],
        fontName="TimesNewRoman",
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
    )

    elements = [Paragraph(date or "", italic_style), Spacer(1, 62), Paragraph("Тапсырыс", title_style), Spacer(1, 12)]

    # --- HEADER ---

    header_text = (
        "«Мемлекеттік әлеуметтік сақтандыру қоры» акционерлік қоғамының "
        "Бухгалтерлік есеп, есептілік және инвестициялық талдау департаменті"
        "<br/>Мемлекеттік корпорацияның шотына келесі артық (қате) төленген "
        "әлеуметтік аударымдарды және (немесе) өсімпұлдар сомаларын аударсын:"
    )

    elements.append(Paragraph(header_text, center_style))
    elements.append(Spacer(1, 26))

    # --- DATA ---
    d5 = e5 = f5 = g5 = h5 = i5 = j5 = k5 = 0
    d6 = e6 = f6 = g6 = h6 = i6 = j6 = k6 = 0

    if rows:
        for amount, count, knp, typ in rows:
            amount = amount or 0
            count = count or 0

            if knp == "026" and typ == "СЗ":
                j5 = count
                k5 = amount
            elif knp == "094" and typ == "СЗ":
                j6 = count
                k6 = amount
            elif knp == "026":
                d5 = count
                e5 = amount
            elif knp == "094" and typ is None:
                f6 = count
                g6 = amount
            elif knp == "094" and typ == "О":
                h6 = count
                i6 = amount

    c5 = e5 + g5 + i5 + k5
    c6 = e6 + g6 + i6 + k6

    b5 = d5 + f5 + h5 + j5
    b6 = d6 + f6 + h6 + j6

    b7 = b5 + b6
    c7 = c5 + c6

    data = [
        ["ТТК", "Барлығы", "", "Соның ішінде", "", "", "", "", "", "", ""],
        ["", "Адам саны", "Сома (теңге)",
         "026 ТТК б-ша ӘА", "",
         "094 ТТК б-ша ӘА өсімпұлы", "",
         "БТ", "",
         "ӨЖҚ", ""],
        ["", "", "",
         "Адам саны", "Сома (теңге)",
         "Адам саны", "Сома (теңге)",
         "Адам саны", "Сома (теңге)",
         "Адам саны", "Сома (теңге)"],

        ["026", b5, f"{c5:,.2f}", d5, f"{e5:,.2f}", f5, f"{g5:,.2f}", h5, f"{i5:,.2f}", j5, f"{k5:,.2f}"],
        ["094", b6, f"{c6:,.2f}", d6, f"{e6:,.2f}", f6, f"{g6:,.2f}", h6, f"{i6:,.2f}", j6, f"{k6:,.2f}"],
        ["", b7, f"{c7:,.2f}", "", "", "", "", "", "", "", ""],
    ]

    table = Table(
        data,
        repeatRows=1,
        colWidths=[40, 60, 80, 60, 80, 60, 80, 60, 80, 60, 80]
    )

    table.setStyle(TableStyle([
        # --- ОБЪЕДИНЕНИЯ ---
        ("SPAN", (0, 0), (0, 2)),  # ТТК
        ("SPAN", (1, 0), (2, 0)),  # Барлығы
        ("SPAN", (3, 0), (10, 0)),  # Соның ішінде

        ("SPAN", (3, 1), (4, 1)),  # 026
        ("SPAN", (5, 1), (6, 1)),  # 094
        ("SPAN", (7, 1), (8, 1)),  # БТ
        ("SPAN", (9, 1), (10, 1)),  # ӨЖҚ

        ("SPAN", (1, 1), (1, 2)),  # Адам саны (Барлығы)
        ("SPAN", (2, 1), (2, 2)),  # Сома (Барлығы)

        # --- СТИЛЬ ---
        ("FONTNAME", (0, 0), (-1, -1), "TimesNewRoman"),
        ("FONTNAME", (0, 0), (-1, 2), "TimesNewRoman-Bold"),

        ("FONTSIZE", (0, 0), (-1, -1), 10),

        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("BACKGROUND", (0, 0), (-1, 2), colors.lightgrey),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 26))

    elements.append(
        Paragraph(
            f'«МӘСҚ» АҚ Басқарма Төрағасының орынбасары: __________________ {approved_by.get("fio")}',
            title_style
        )
    )

    elements.append(Spacer(1, 62))
    elements.append(Paragraph(f"Орынд. {fio or ''}", italic_style))

    doc.build(elements)

    buffer.seek(0)
    return buffer


def rows_to_excel(rows=None, date=None, fio=None, approved_by=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    offset = 10

    def r(row):
        return row + offset

    def cell(col, row):
        return f"{col}{r(row)}"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    bold = Font(bold=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # -------- TOP TEXT (HEADER OF DOCUMENT) --------

    ws.merge_cells("A6:K6")
    ws.merge_cells("A8:K8")

    ws["K1"] = f"{date}"
    ws["K1"].alignment = right
    ws["I1"].font = Font(italic=True)

    ws["A6"] = "Тапсырыс"
    ws["A6"].alignment = center
    ws["A6"].font = Font(bold=True, size=14)

    ws["A8"] = (
        "«Мемлекеттік әлеуметтік сақтандыру қоры» акционерлік қоғамының Бухгалтерлік есеп, есептілік және инвестициялық талдау департаменті"
        "\nМемлекеттік корпорацияның шотына келесі артық (қате) төленген әлеуметтік аударымдарды және (немесе) өсімпұлдар сомаларын аударсын:"
    )
    ws["A8"].alignment = center
    ws.row_dimensions[8].height = 60

    # -------- TABLE MERGES --------

    merges = [
        ("A",1,"A",4), ("B",1,"C",1), ("D",1,"K",1),
        ("B",2,"B",4), ("C",2,"C",4),
        ("D",2,"E",2), ("F",2,"G",2), ("H",2,"I",2),
        ("J", 2, "K", 2),
        ("J", 3, "J", 4), ("K", 3, "K", 4),
        ("D",3,"D",4), ("E",3,"E",4), ("F",3,"F",4),
        ("G",3,"G",4), ("H",3,"H",4), ("I",3,"I",4)
    ]

    for c1, r1, c2, r2 in merges:
        ws.merge_cells(f"{c1}{r(r1)}:{c2}{r(r2)}")

    # -------- TABLE HEADER TEXT --------

    header = {
        ("A",1): "ТТК",
        ("B",1): "Барлығы",
        ("B",2): "Адам саны",
        ("C",2): "Сома (теңге)",
        ("D",1): "Соның ішінде",
        ("D",2): "026 ТТК б-ша ƏА",
        ("F",2): "094 ТТК б-ша ƏА өсімпұлы",
        ("H",2): "БТ",
        ("D",3): "Адам саны",
        ("E",3): "Сома (теңге)",
        ("F",3): "Адам саны",
        ("G",3): "Сома (теңге)",
        ("H",3): "Адам саны",
        ("I",3): "Сома (теңге)",
        ("J", 2): "ӨЖҚ",
        ("J", 3): "Адам саны",
        ("K", 3): "Сома (теңге)",
    }

    for (col, row), text in header.items():
        ws[cell(col, row)] = text

    ws[cell("A",5)] = "026"
    ws[cell("A",6)] = "094"

    # -------- STYLES --------

    for row in ws[f"A{r(1)}:K{r(4)}"]:
        for c in row:
            c.alignment = center
            c.font = bold
            c.border = border

    for row in ws[f"A{r(5)}:K{r(7)}"]:
        for c in row:
            c.alignment = center
            c.border = border

    # -------- NUMBER FORMAT --------

    for col, row in [
        ("E",5), ("G",5), ("I",5),
        ("E",6), ("G",6), ("I",6),
        ("C",5), ("C",6), ("C",7),
        ("K", 5), ("K", 6)
    ]:
        ws[cell(col, row)].number_format = '# ### ### ##0.00'

    # -------- DATA --------

    if rows:
        for amount, count, knp, typ in rows:

            if knp == "026" and typ == "СЗ":
                ws[cell("J", 5)] = count
                ws[cell("K", 5)] = amount
            elif knp == "094" and typ == "СЗ":
                ws[cell("J", 6)] = count
                ws[cell("K", 6)] = amount
            elif knp == "026":
                ws[cell("D",5)] = count
                ws[cell("E",5)] = amount
            elif knp == "094" and typ is None:
                ws[cell("F",6)] = count
                ws[cell("G",6)] = amount
            elif knp == "094" and typ == "О":
                ws[cell("H",6)] = count
                ws[cell("I",6)] = amount

    for row in ws[f"D{r(5)}:K{r(6)}"]:
        for c in row:
            if c.value is None:
                c.value = 0
    # -------- SUM FUNCTION --------

    def s(coords):
        return sum(ws[cell(c, r)].value or 0 for c, r in coords)

    ws[cell("C", 5)] = s([("E", 5), ("G", 5), ("I", 5), ("K", 5)])
    ws[cell("C", 6)] = s([("E", 6), ("G", 6), ("I", 6), ("K", 6)])

    ws[cell("B", 5)] = s([("D", 5), ("F", 5), ("H", 5), ("J", 5)])
    ws[cell("B", 6)] = s([("D", 6), ("F", 6), ("H", 6), ("J", 6)])

    ws[cell("C", 7)] = s([
        ("E", 5), ("G", 5), ("I", 5), ("K", 5),
        ("E", 6), ("G", 6), ("I", 6), ("K", 6)
    ])

    ws[cell("B", 7)] = s([
        ("D", 5), ("F", 5), ("H", 5), ("J", 5),
        ("D", 6), ("F", 6), ("H", 6), ("J", 6)
    ])

    ws.merge_cells(f"{cell('A',10)}:{cell('K',10)}")
    ws[cell("A",10)] = f'«МӘСҚ» АҚ Басқарма Төрағасының орынбасары: __________________ {approved_by.get("fio")}'
    ws[cell("A",10)].alignment = center
    ws[cell("A",10)].font = Font(bold=True, size=14)

    ws.merge_cells(f"{cell('A', 15)}:{cell('K', 15)}")
    ws[cell("A", 15)] = f"Орынд. {fio}"
    ws[cell("A", 15)].alignment = left
    ws[cell("A", 15)].font = Font(italic=True)

    # -------- COLUMN WIDTH --------

    widths = [5, 10, 15, 10, 15, 10, 15, 10, 15, 10, 15]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream